# Program that writes codeset values to a table in Microsoft Excel
# Used for some analysis work for CalcGuard Technologies, Inc.
# Figured it might be an example in accessing the repository.
# Jim Northey
# import pytest
from pyfixorchestra import FixDictionary
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

def writeFields(ws,fields):
    ws.append(["Fieldname","Tag","Datatype","Description"])
    rows=1
    for field in fields:
        ws.append([field['name'],field['id'],field['type'],field['documentation'][0].get("text"),""])
        rows+=1
        print(field)
    fieldtable=Table(displayName="FieldsTable", ref=f"A1:D{rows}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                      showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    fieldtable.tableStyleInfo = style
    ws.add_table(fieldtable)
    return   

def writeDatatypes(ws,datatypes):
    ws.append(["FIXDatatype","Description","Datatype"])
    rows=1
    for datatype in datatypes:
        ws.append([datatype['name'],datatype['documentation'][0].get("text"),""])
        rows+=1
    datatypetable=Table(displayName="DataTypeTable", ref=f"A1:C{rows}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                      showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    datatypetable.tableStyleInfo = style
    ws.add_table(datatypetable)
    return




def writeCodesetTypes(ws,codesets):
    ws.append(["CodeSetName","CodeSetExternalID","Datatype"])
    rows=1
    for codeset in codesets:
        ws.append([codeset['name'],codeset['id'],codeset['type']])
        rows+=1
    codetable=Table(displayName="CodeSetTypeTable", ref=f"A1:C{rows}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                      showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    codetable.tableStyleInfo = style
    ws.add_table(codetable)
    return



def writeCodesets(ws,codesets):
    ws.append(["CodeSetName","CodeSetExternalID","Datatype","Code","Value","Mneumonic","Description","Mapping","CodeSetSource","CodeSetVersion"])
    rows=1
    for codeset in codesets:
        #print(codeset)
        for codes in codeset['codes']:
            elaboration=''
            desc=''
            for text in codes['documentation']:
                if(text['purpose']=="SYNOPSIS"):
                    desc=text['text']
                elif(text['purpose']=='ELABORATION'):
                    elaboration=text['text']
            ws.append([codeset['name'],codeset['id'],codeset['type'],codes['value'],desc,codes['name'],elaboration,"","FIX Trading Community","FIXLatest"])
            rows+=1
            #print (codes)
    codetable=Table(displayName="CodeSetTable", ref=f"A1:J{rows}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                      showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    codetable.tableStyleInfo = style
    ws.add_table(codetable)
    return



fd = FixDictionary()
fd.read_xml("resource_dir/OrchestraFIXLatest.xml")
codesets=fd.generateDictionary("codeSets")
datatypes=fd.generateDictionary("datatypes")
fields=fd.generateDictionary("fields")
wb=openpyxl.Workbook()
ws=wb.active
ws.title="Codesets"
writeCodesets(ws,codesets)
ws=wb.create_sheet("CodeSetTypes")
writeCodesetTypes(ws,codesets)
ws=wb.create_sheet("Datatypes")
writeDatatypes(ws,datatypes)
ws=wb.create_sheet("Fields")
writeFields(ws,fields)
wb.save("FIX_metadata.xlsx")



